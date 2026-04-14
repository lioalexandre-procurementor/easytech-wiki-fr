#!/usr/bin/env python3
"""Génère le tracker de vérification émulateur (XLSX)."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs" / "WC4-Verification-Tracker.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

units = json.loads((ROOT / "data/wc4/elite-units/_index.json").read_text())
gens = json.loads((ROOT / "data/wc4/generals/_index.json").read_text())

# ------- Styles -------
FONT = "Arial"
HEAD = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=14, color="1a2230")
SUB = Font(name=FONT, italic=True, size=10, color="5a6878")
HEAD_FILL = PatternFill("solid", start_color="1a2230")
BAND_FILL = PatternFill("solid", start_color="f4f1e8")
SCORPION_FILL = PatternFill("solid", start_color="fde8e6")
STANDARD_FILL = PatternFill("solid", start_color="fdf6e3")
TODO_FILL = PatternFill("solid", start_color="ffe58f")
DONE_FILL = PatternFill("solid", start_color="c2f0c2")
THIN = Side(border_style="thin", color="c8c8c8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ============================================================================
# README sheet
# ============================================================================
readme = wb.active
readme.title = "README"
readme["A1"] = "WC4 — Tracker de vérification in-game"
readme["A1"].font = TITLE
readme["A2"] = "Wiki : easytech-wiki (Alex)"
readme["A2"].font = SUB

readme["A4"] = "Comment l'utiliser"
readme["A4"].font = Font(name=FONT, bold=True, size=12)

instructions = [
    "",
    "1. Lance l'émulateur (LDPlayer recommandé) avec World Conqueror 4.",
    "2. Ouvre l'onglet « Units » ou « Generals » de ce tracker.",
    "3. Pour chaque ligne, fais une capture d'écran dans l'émulateur au menu concerné.",
    "4. Coche la colonne correspondante en mettant « OK » (ou ✓). Toute autre valeur = TODO.",
    "5. Colle les valeurs numériques dans les colonnes dédiées (Level 1 ATK, etc.).",
    "6. Remplis la colonne « Notes » si quelque chose diverge du wiki actuel.",
    "7. Quand toutes les colonnes d'une ligne sont OK, mets la colonne « Verified » à OK.",
    "8. Quand tu as fini un lot, envoie-moi le tracker — je pousse les corrections dans le JSON.",
    "",
    "Conventions de couleur :",
    "  • Jaune = TODO (valeur manquante)",
    "  • Vert = OK",
    "  • Rouge clair = ligne Scorpion (campagne, obtention différente)",
    "  • Crème = ligne Standard",
    "",
    "Passes recommandées (plus efficace que unit-par-unit) :",
    "  A. Elite Unit stats (Level 1, 5, 9, 12) — onglet Units, colonnes L1-L12.",
    "  B. Elite Unit perks / skills unlock levels — onglet Units, colonne Perks OK.",
    "  C. Generals — screenshots des 3 slots + stars + attributs — onglet Generals.",
    "  D. Trained Generals — re-capture post-entraînement si disponible.",
    "  E. Scorpion campagne (Modern War) — ligne colorée rouge.",
    "",
    "Screenshots : range-les dans public/images/units/<slug>.png et",
    "public/images/generals/<slug>.png dans le projet — ça branchera automatiquement.",
]
for i, line in enumerate(instructions, start=5):
    readme.cell(row=i, column=1, value=line).font = BODY

readme.column_dimensions["A"].width = 95

# ============================================================================
# UNITS sheet
# ============================================================================
us = wb.create_sheet("Units")
unit_cols = [
    ("Slug", 18),
    ("Name", 24),
    ("Faction", 10),
    ("Category", 11),
    ("Country", 10),
    ("Tier", 7),
    ("Stats L1 OK", 12),
    ("Stats L5 OK", 12),
    ("Stats L9 OK", 12),
    ("Stats L12 OK", 13),
    ("Perks OK", 11),
    ("Screenshot OK", 15),
    ("ATK L1", 9),
    ("ATK L12", 10),
    ("DEF L1", 9),
    ("DEF L12", 10),
    ("HP L1", 9),
    ("HP L12", 10),
    ("Verified", 11),
    ("Notes", 40),
]
for i, (h, w) in enumerate(unit_cols, start=1):
    c = us.cell(row=1, column=i, value=h)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = CENTER
    c.border = BORDER
    us.column_dimensions[get_column_letter(i)].width = w
us.row_dimensions[1].height = 24
us.freeze_panes = "C2"

# Sort: faction (standard first), category, name
CAT_ORDER = {"tank": 0, "infantry": 1, "artillery": 2, "navy": 3, "airforce": 4}
units_sorted = sorted(
    units,
    key=lambda u: (u["faction"] != "standard", CAT_ORDER.get(u["category"], 99), u["name"]),
)
for r, u in enumerate(units_sorted, start=2):
    row = [
        u["slug"], u["name"], u["faction"], u["category"], u["country"],
        u["tier"], "TODO", "TODO", "TODO", "TODO", "TODO", "TODO",
        "", "", "", "", "", "", "TODO", "",
    ]
    for i, val in enumerate(row, start=1):
        c = us.cell(row=r, column=i, value=val)
        c.font = BODY
        c.border = BORDER
        c.alignment = LEFT if i in (2, 20) else CENTER
    fill = SCORPION_FILL if u["faction"] == "scorpion" else STANDARD_FILL
    for i in range(1, len(row) + 1):
        if i in (7, 8, 9, 10, 11, 12, 19):
            continue  # keep state cells white-ish for CF
        us.cell(row=r, column=i).fill = fill

# Conditional formatting — TODO (yellow) / OK (green) on state columns
state_cols = [7, 8, 9, 10, 11, 12, 19]  # Stats L1, L5, L9, L12, Perks, Screenshot, Verified
for col in state_cols:
    col_letter = get_column_letter(col)
    rng = f"{col_letter}2:{col_letter}{len(units_sorted) + 1}"
    us.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"OK"'], fill=DONE_FILL),
    )
    us.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"TODO"'], fill=TODO_FILL),
    )

# Summary at the top of a column past the data
summary_row = len(units_sorted) + 3
us.cell(row=summary_row, column=1, value="Total:").font = Font(name=FONT, bold=True)
us.cell(row=summary_row, column=2, value=len(units_sorted))
us.cell(row=summary_row + 1, column=1, value="Vérifiés:").font = Font(name=FONT, bold=True)
us.cell(
    row=summary_row + 1, column=2,
    value=f'=COUNTIF(S2:S{len(units_sorted)+1},"OK")'
)
us.cell(row=summary_row + 2, column=1, value="% fait:").font = Font(name=FONT, bold=True)
us.cell(
    row=summary_row + 2, column=2,
    value=f'=B{summary_row+1}/B{summary_row}'
)
us.cell(row=summary_row + 2, column=2).number_format = "0.0%"

# ============================================================================
# GENERALS sheet
# ============================================================================
gs = wb.create_sheet("Generals")
gen_cols = [
    ("Slug", 16),
    ("Name", 24),
    ("Faction", 10),
    ("Category", 12),
    ("Country", 10),
    ("Rank", 7),
    ("Has trained?", 13),
    ("Skill 1 OK", 11),
    ("Skill 2 OK", 11),
    ("Skill 3 OK", 11),
    ("Skill stars OK", 14),
    ("Attributes OK", 13),
    ("Trained form OK", 15),
    ("Acquisition type", 15),
    ("Cost", 10),
    ("Currency", 11),
    ("Portrait OK", 12),
    ("Verified", 10),
    ("Notes", 40),
]
for i, (h, w) in enumerate(gen_cols, start=1):
    c = gs.cell(row=1, column=i, value=h)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = CENTER
    c.border = BORDER
    gs.column_dimensions[get_column_letter(i)].width = w
gs.row_dimensions[1].height = 24
gs.freeze_panes = "C2"

gens_sorted = sorted(
    gens,
    key=lambda g: (g["faction"] != "standard", g["rank"], g["name"]),
)
for r, g in enumerate(gens_sorted, start=2):
    has_trained = "OUI" if g.get("hasTrained") else "non"
    trained_col = "TODO" if g.get("hasTrained") else "N/A"
    row = [
        g["slug"], g["name"], g["faction"], g["category"], g["country"], g["rank"],
        has_trained,
        "TODO", "TODO", "TODO", "TODO", "TODO", trained_col,
        g.get("acquisition", ""), "", "",
        "TODO", "TODO", "",
    ]
    for i, val in enumerate(row, start=1):
        c = gs.cell(row=r, column=i, value=val)
        c.font = BODY
        c.border = BORDER
        c.alignment = LEFT if i in (2, 19) else CENTER
    fill = SCORPION_FILL if g["faction"] == "scorpion" else STANDARD_FILL
    for i in range(1, len(row) + 1):
        if i in (8, 9, 10, 11, 12, 13, 17, 18):
            continue
        gs.cell(row=r, column=i).fill = fill

gen_state_cols = [8, 9, 10, 11, 12, 13, 17, 18]
for col in gen_state_cols:
    col_letter = get_column_letter(col)
    rng = f"{col_letter}2:{col_letter}{len(gens_sorted) + 1}"
    gs.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OK"'], fill=DONE_FILL))
    gs.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"TODO"'], fill=TODO_FILL))

g_summary = len(gens_sorted) + 3
gs.cell(row=g_summary, column=1, value="Total:").font = Font(name=FONT, bold=True)
gs.cell(row=g_summary, column=2, value=len(gens_sorted))
gs.cell(row=g_summary + 1, column=1, value="Vérifiés:").font = Font(name=FONT, bold=True)
gs.cell(row=g_summary + 1, column=2, value=f'=COUNTIF(R2:R{len(gens_sorted)+1},"OK")')
gs.cell(row=g_summary + 2, column=1, value="% fait:").font = Font(name=FONT, bold=True)
gs.cell(row=g_summary + 2, column=2, value=f'=B{g_summary+1}/B{g_summary}')
gs.cell(row=g_summary + 2, column=2).number_format = "0.0%"

# ============================================================================
# ROSTER BACKLOG sheet — generals à ajouter (phase 2+)
# ============================================================================
rb = wb.create_sheet("Roster Backlog")
rb["A1"] = "Généraux à ajouter au wiki (Phase 2 → 4)"
rb["A1"].font = TITLE
rb["A2"] = "Source : tier lists Fandom/NamuWiki — à vérifier et compléter avec l'émulateur."
rb["A2"].font = SUB

backlog_head = [("Name", 26), ("Country", 12), ("Suggested category", 20),
                ("Suggested rank", 14), ("Phase", 9), ("In wiki?", 11), ("Notes", 40)]
for i, (h, w) in enumerate(backlog_head, start=1):
    c = rb.cell(row=4, column=i, value=h)
    c.font = HEAD
    c.fill = HEAD_FILL
    c.alignment = CENTER
    c.border = BORDER
    rb.column_dimensions[get_column_letter(i)].width = w
rb.row_dimensions[4].height = 22

backlog = [
    # Phase 2 — S/A tier to capture
    ("Erich von Manstein", "DE", "tank", "S", 2),
    ("Walther Model", "DE", "tank", "A", 2),
    ("Gerd von Rundstedt", "DE", "infantry", "A", 2),
    ("Albert Kesselring", "DE", "airforce", "A", 2),
    ("Hermann Goering", "DE", "airforce", "B", 3),
    ("Erich Raeder", "DE", "navy", "A", 2),
    ("Omar Bradley", "US", "infantry", "A", 2),
    ("Douglas MacArthur", "US", "balanced", "A", 2),
    ("Dwight Eisenhower", "US", "balanced", "S", 2),
    ("Chester Nimitz", "US", "navy", "S", 2),
    ("William Halsey", "US", "navy", "A", 2),
    ("Raymond Spruance", "US", "navy", "A", 2),
    ("Curtis LeMay", "US", "airforce", "A", 2),
    ("Isoroku Yamamoto", "JP", "navy", "S", 2),
    ("Chuichi Nagumo", "JP", "navy", "A", 2),
    ("Minoru Genda", "JP", "airforce", "A", 2),
    ("Hideki Tojo", "JP", "balanced", "B", 3),
    ("Tomoyuki Yamashita", "JP", "infantry", "A", 2),
    ("Vasily Chuikov", "RU", "infantry", "A", 2),
    ("Semyon Timoshenko", "RU", "tank", "A", 2),
    ("Aleksandr Vasilevsky", "RU", "balanced", "A", 2),
    ("Rodion Malinovsky", "RU", "tank", "A", 2),
    ("Andrey Yeryomenko", "RU", "balanced", "B", 3),
    ("Harold Alexander", "GB", "balanced", "A", 2),
    ("Archibald Wavell", "GB", "infantry", "B", 3),
    ("William Slim", "GB", "infantry", "A", 2),
    ("Andrew Cunningham", "GB", "navy", "A", 2),
    ("Arthur Harris", "GB", "airforce", "A", 2),
    ("Hugh Dowding", "GB", "airforce", "A", 2),
    ("Charles de Gaulle", "FR", "balanced", "B", 2),
    ("Philippe Leclerc", "FR", "tank", "B", 3),
    ("Jean de Lattre de Tassigny", "FR", "infantry", "B", 3),
    ("Pietro Badoglio", "IT", "balanced", "B", 3),
    ("Italo Balbo", "IT", "airforce", "B", 3),
    ("Ugo Cavallero", "IT", "balanced", "C", 4),
    ("Władysław Sikorski", "PL", "balanced", "B", 3),
    ("Władysław Anders", "PL", "infantry", "B", 3),
    ("Chiang Kai-shek", "CN", "balanced", "A", 2),
    ("Zhu De", "CN", "infantry", "B", 3),
    ("Lin Biao", "CN", "infantry", "A", 2),
    ("Peng Dehuai", "CN", "infantry", "A", 2),
    ("Moshe Dayan", "IL", "tank", "A", 2),
    ("David Ben-Gurion", "IL", "balanced", "B", 3),
    ("Norman Schwarzkopf", "US", "tank", "S", 2),
    ("Colin Powell", "US", "balanced", "A", 2),
    ("Creighton Abrams", "US", "tank", "A", 2),
]
for i, (name, country, cat, rank, phase) in enumerate(backlog, start=5):
    data = [name, country, cat, rank, f"P{phase}", "NON", ""]
    for j, v in enumerate(data, start=1):
        c = rb.cell(row=i, column=j, value=v)
        c.font = BODY
        c.border = BORDER
        c.alignment = LEFT if j in (1, 7) else CENTER

rb.cell(row=5 + len(backlog) + 1, column=1,
        value=f"{len(backlog)} généraux candidats identifiés — liste non exhaustive.").font = SUB

wb.save(str(OUT))
print(f"✓ Tracker saved to {OUT}")
print(f"  Units: {len(units_sorted)} | Generals: {len(gens_sorted)} | Backlog: {len(backlog)}")
